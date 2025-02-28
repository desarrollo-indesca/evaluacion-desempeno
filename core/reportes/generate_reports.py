
import openpyxl
import datetime
from evaluacion.models import Evaluacion
from io import BytesIO
from django.http import HttpResponse

def create_dnf(periodo, file_path='core/reportes/bases/plan-anual-formacion.xlsx'):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Access the first worksheet
    worksheet = workbook.active

    # Write data to the worksheet
    # Fetch data from the database
    evaluaciones = Evaluacion.objects.filter(periodo=periodo, estado="A").order_by(
        'evaluado__user__first_name'
    )
    
    # Write data to the worksheet
    row = 10
    i = 1
    for evaluacion in evaluaciones:
        for formacion in enumerate(evaluacion.formaciones.filter(anadido_por='H'), start=1):            
            formacion_sugerida = formacion.necesidad_formacion.upper()
            formacion_especifica = ''
            ente_didactico = ''
            instructor = ''
            duracion_horas = ''
            ano = datetime.datetime.now().year
            fecha_planificada = ''
            fecha_real = ''
            prioridad = formacion.prioridad

            # Write data to the worksheet
            data = [
                i,evaluacion.evaluado.ficha, evaluacion.evaluado.user.get_full_name(), 
                evaluacion.evaluado.cargo.nombre.upper(), evaluacion.evaluado.gerencia.nombre.upper(), 
                evaluacion.evaluado.tipo_personal.nombre.upper(),
                formacion_sugerida, formacion_especifica, ente_didactico,
                instructor, duracion_horas, ano, fecha_planificada,
                fecha_real, prioridad, '', '', '',
                'X' if formacion.competencias.filter(nombre__iexact='CAPACIDADES OPERATIVAS').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='CAPACIDADES ANALÍTICAS Y DE SÍNTESIS').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Capacidades de Negociación y Relación').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Capacidades de Gestión').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Desarrollo').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Liderazgo').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Adaptabilidad').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Realización de Tareas').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Producción').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Desarrollo de los demás').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Desarrollo Personal').exists() else '',
                'X' if formacion.competencias.filter(nombre__iexact='Comunicación').exists() else '',
            ]
            
            for col_num, cell_data in enumerate(data, start=1):
                worksheet.cell(row=row, column=col_num, value=cell_data)

            i += 1            
            row += 1
    
    # Save the workbook to a file
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # add the picture in static/img/LogoDeIndesca.svg in A3
    img = openpyxl.drawing.image.Image('static/img/LogoDeIndesca.png')
    img.anchor = 'A3'
    worksheet.add_image(img)
    
    # Return the report as a file response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="dnf_{periodo.fecha_inicio.strftime("%Y-%m-%d")}_{periodo.fecha_fin.strftime("%Y-%m-%d")}.xlsx"'
    return response

def obtener_accion_a_tomar_desempeno(escala):
    if(escala <= 2):
        return "2.00 - Observación directa máximo de seis meses."
    elif(escala > 2 and escala <= 2.75):
        return "2.01 a 2.75 - Observación de un año."
    elif(escala > 2.75 and escala <= 3.5):
        return "2.76 a 3.50 - Ofrecerle oportunidades  de crecimiento  dentro  de su cargo o área de trabajo."
    elif(escala > 3.5 and escala <= 4.25):
        return "3.51 a 4.25 - Califica para  promociones  (tanto en el escalafón como en la estructura administrativa,  previo cumplimiento de todos los requisitos del cargo o nivel)."
    else:
        return "4.26 hasta 5 - Califica para promociones (tanto en el escalafón como en la estructura  administrativa,  previo  cumplimiento de todos los requisitos del cargo o nivel)."

def obtener_accion_a_tomar_ct_operativos(escala):
    if(escala <= 2.5):
        return "< 2,50 - Cerrar brecha existente a través de formación.  Se requiere de una observación directa por parte del supervisor."
    elif(escala > 2.5 and escala <= 2.99):
        return "2,51 a 2,99 - Cerrar brecha existente a través de formación.  Se requiere de una observación directa por parte del supervisor."
    elif(escala > 2.99 and escala <= 3.49):
        return "2,00 a 3,49 - Cerrar brecha existente a través de formación.  "
    elif(escala > 3.49 and escala <= 4.50):
        return "3,50 a 4,50 - Cerrar brecha existente a través de formación.  "
    else:
        return "4,51 hasta 5,00 - El haber alcanzado este puntaje  no significa que no tenga brechas ya que cada nivel es distinto y  las prioridades  serán  subsanar  los aspectos  más rezagados, ya que no todo puede resolverse en un solo periodo.  Se observa que mantenga su posición alcanzada."

def obtener_accion_a_tomar_ct_apoyo(escala):
    if(escala <= 2.5):
        return "< 2,50 - Cerrar brecha existente a través de formación.  Se requiere de una observación directa por parte del supervisor."
    elif(escala > 2.5 and escala <= 2.99):
        return "2,51 a 2,99 - Cerrar brecha existente a través de formación.  Se requiere de una observación directa por parte del supervisor."
    elif(escala > 2.99 and escala <= 3.49):
        return "2,00 a 3,49 - Cerrar brecha existente a través de formación."
    elif(escala > 3.49 and escala <= 4.50):
        return "3,50 a 4,50 - Cerrar brecha existente a través de formación."
    else:
        return "4,51 hasta 5,00 - Si  la  brecha  se  cerró  no  se  tomarán   acciones. Se observa  que mantenga su posición alcanzado."

def obtener_accion_a_tomar_genericas(escala):
    if(escala <= 2.5):
        return "< 2,50 - Cerrar brecha existente a través de formación."
    elif(escala > 2.5 and escala <= 2.99):
        return "2,51 a 2,99 - Cerrar brecha existente a través de formación."
    elif(escala > 2.99 and escala <= 3.49):
        return "2,00 a 3,49 - Cerrar brecha existente a través de formación."
    elif(escala > 3.49 and escala <= 4.50):
        return "3,50 a 4,50 - Cerrar brecha existente a través de formación."
    else:
        return "4,51 hasta 5,00 - Si la brecha se cerró no se tomarán acciones."

def fill_resumen_periodo(periodo):
    # fetch all the evaluations of the period
    evaluaciones = Evaluacion.objects.filter(
        periodo=periodo,
        estado='A'
    ).order_by('evaluado__user__first_name')
        
    # Load the Excel workbook
    workbook = openpyxl.load_workbook('core/reportes/bases/resumen.xlsx')

    # Access the first worksheet
    worksheet = workbook.active
    
    # Fill the worksheet with the resumen data
    row = 8
    for i,evaluacion in enumerate(evaluaciones, start=1):
        datos_personal = evaluacion.evaluado

        worksheet.cell(row=row, column=1, value=i)
        worksheet.cell(row=row, column=2, value=datos_personal.ficha)
        worksheet.cell(row=row, column=3, value=datos_personal.user.get_full_name())
        worksheet.cell(row=row, column=4, value=datos_personal.fecha_ingreso.strftime("%d/%m/%Y"))
        worksheet.cell(row=row, column=5, value=(datetime.date.today() - datos_personal.fecha_ingreso).days // 365)
        worksheet.cell(row=row, column=6, value=datos_personal.tipo_personal.nombre)
        worksheet.cell(row=row, column=7, value=datos_personal.gerencia.nombre)
        worksheet.cell(row=row, column=8, value=datos_personal.cargo.nombre)
        worksheet.cell(row=row, column=9, value=datos_personal.escalafon.nivel)
        
        resultado_instrumento = evaluacion.resultados.filter(instrumento__nombre__icontains='Evaluación del Desempeño').first()
        worksheet.cell(row=row, column=10, value=resultado_instrumento.resultado_final if resultado_instrumento else '')
        worksheet.cell(row=row, column=11, value=obtener_accion_a_tomar_desempeno(resultado_instrumento.resultado_final) if resultado_instrumento else '')
        
        resultado_ct_tecnicas = evaluacion.resultados.filter(instrumento__nombre__icontains='Competencias Técnicas').first()
        worksheet.cell(row=row, column=12, value=resultado_ct_tecnicas.resultado_final if resultado_ct_tecnicas else '')
        
        tipo_personal = datos_personal.tipo_personal
        if tipo_personal.nombre == 'APOYO':
            worksheet.cell(row=row, column=13, value=obtener_accion_a_tomar_ct_apoyo(resultado_ct_tecnicas.resultado_final) if resultado_ct_tecnicas else '')
        else:
            worksheet.cell(row=row, column=13, value=obtener_accion_a_tomar_ct_operativos(resultado_ct_tecnicas.resultado_final) if resultado_ct_tecnicas else '')
        
        resultado_ct_genericas = evaluacion.resultados.filter(instrumento__nombre__icontains='Competencias Genérica').first()
        worksheet.cell(row=row, column=14, value=resultado_ct_genericas.resultado_final if resultado_ct_genericas else '')
        worksheet.cell(row=row, column=15, value=obtener_accion_a_tomar_genericas(resultado_ct_genericas.resultado_final) if resultado_ct_genericas else '')
        
        row += 1
    
    # Save the workbook to a file
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # add the picture in static/img/LogoDeIndesca.png
    img = openpyxl.drawing.image.Image('static/img/LogoDeIndesca.png')
    img.anchor = 'A3'
    worksheet.add_image(img)
    
    # Return the report as a file response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="resumen_{periodo.fecha_inicio.strftime("%Y-%m-%d")}_{periodo.fecha_fin.strftime("%Y-%m-%d")}.xlsx"'
    return response

def fill_resultado_operativo(evaluacion):
    workbook = openpyxl.load_workbook('core/reportes/bases/resultados-operativos.xlsx')
    worksheet = workbook.active

    worksheet.cell(row=6, column=1, value=f"EVALUACIÓN DE DESEMPEÑO")
    worksheet.cell(row=6, column=1, value=f"PERÍODO DE EVALUACIÓN DEL {evaluacion.periodo.fecha_inicio.strftime('%d/%m/%Y')} AL {evaluacion.periodo.fecha_fin.strftime('%d/%m/%Y')}")

    worksheet.cell(row=14, column=7, value=evaluacion.escalafones.get(asignado_por="H").escalafon.nivel.upper())

    worksheet.cell(row=8, column=2, value=evaluacion.evaluado.user.get_full_name())
    worksheet.cell(row=8, column=18, value=evaluacion.evaluado.ficha)
    worksheet.cell(row=9, column=2, value=evaluacion.evaluado.cargo.nombre)
    
    resultado_instrumento = evaluacion.resultados.get(instrumento__nombre__icontains='Evaluación del Desempeño')
    worksheet.cell(row=13, column=2, value=float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Costos').resultado_final) / float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Costos').seccion.peso) * (float(resultado_instrumento.instrumento.peso) * float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Costos').seccion.peso) / 100))
    worksheet.cell(row=13, column=3, value=float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Cantidad').resultado_final) / float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Cantidad').seccion.peso) * (float(resultado_instrumento.instrumento.peso) * float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Cantidad').seccion.peso) / 100))
    worksheet.cell(row=13, column=4, value=float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Calidad').resultado_final) / float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Calidad').seccion.peso) * (float(resultado_instrumento.instrumento.peso) * float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Calidad').seccion.peso) / 100))
    worksheet.cell(row=13, column=5, value=float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Seguridad').resultado_final) / float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Seguridad').seccion.peso) * (float(resultado_instrumento.instrumento.peso) * float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Seguridad').seccion.peso) / 100))
    worksheet.cell(row=13, column=6, value=float(resultado_instrumento.resultado_final))
    
    resultado_instrumento = evaluacion.resultados.get(instrumento__nombre__icontains='Competencias Técnicas')
    worksheet.cell(row=13, column=7, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Capacidades Operativas').resultado_final)
    worksheet.cell(row=13, column=8, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Capacidades Analíticas y de Síntesis').resultado_final)
    worksheet.cell(row=13, column=9, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Capacidades de Negociación y Relación').resultado_final)
    worksheet.cell(row=13, column=10, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Capacidades de Gestión').resultado_final)
    worksheet.cell(row=13, column=11, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Desarrollo').resultado_final)
    worksheet.cell(row=13, column=12, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Otros').resultado_final)
    worksheet.cell(row=13, column=13, value=resultado_instrumento.resultado_final)

    resultado_instrumento = evaluacion.resultados.get(instrumento__nombre__icontains='Competencias Genérica')
    worksheet.cell(row=13, column=14, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Liderazgo:').resultado_final)
    worksheet.cell(row=13, column=15, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Adaptabilidad:').resultado_final)
    worksheet.cell(row=13, column=16, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Relaciones:').resultado_final)
    worksheet.cell(row=13, column=17, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Realización de Tareas:').resultado_final)
    worksheet.cell(row=13, column=18, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Producción:').resultado_final)
    worksheet.cell(row=13, column=19, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Desarrollo de los demás:').resultado_final)
    worksheet.cell(row=13, column=20, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Desarrollo Personal:').resultado_final)
    worksheet.cell(row=13, column=21, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Comunicación:').resultado_final)
    worksheet.cell(row=13, column=22, value=resultado_instrumento.resultado_final)

    logros = evaluacion.logros_y_metas.filter(anadido_por="H", periodo="A")
    formaciones = evaluacion.formaciones.filter(anadido_por="H")

    tecnico_cientificas = formaciones.filter(clasificacion__clasificacion__icontains='Tecno-Cientifica').count()
    supervisoria_formativa = formaciones.filter(clasificacion__clasificacion__icontains='Supervisoria Formativa').count()
    tecnica_administrativa = formaciones.filter(clasificacion__clasificacion__icontains='Tecnica Administrativa').count()
    seguridad_industrial = formaciones.filter(clasificacion__clasificacion__icontains='Seguridad Industrial').count()
    total_cursos = formaciones.count()

    worksheet.cell(row=40, column=3, value=tecnico_cientificas)
    worksheet.cell(row=41, column=3, value=supervisoria_formativa)
    worksheet.cell(row=40, column=12, value=tecnica_administrativa)
    worksheet.cell(row=40, column=12, value=seguridad_industrial)
    worksheet.cell(row=40, column=21, value=total_cursos)
    for i, logro in enumerate(logros, start=17):
        worksheet.cell(row=i, column=1, value=logro.descripcion)
        worksheet.cell(row=i, column=16, value=logro.porc_cumplimiento)

    for i, formacion in enumerate(formaciones, start=31):
        worksheet.cell(row=i, column=1, value=formacion.necesidad_formacion)
        worksheet.cell(row=i, column=5, value=formacion.prioridad)
        worksheet.cell(row=i, column=7, value=formacion.clasificacion.clasificacion)
        col = 13
        if formacion.competencias.filter(nombre__icontains='Capacidades Operativas').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Capacidades de Negociación y Relación').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Capacidades de Gestión').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Desarrollo').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Liderazgo').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Adaptabilidad').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Relaciones').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Realización de Tareas').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Producción').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Desarrollo de los demás').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Desarrollo del Personal').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Comunicación').exists():
            worksheet.cell(row=i, column=col, value='X')

    metas = evaluacion.logros_y_metas.filter(periodo="P", anadido_por="H")[:5]
    for i, meta in enumerate(metas, start=44):
        worksheet.cell(row=i, column=1, value=meta.descripcion)
        worksheet.cell(row=i, column=16, value=meta.prioridad_larga())

    if evaluacion.comentario_evaluado:
        worksheet.cell(row=53, column=1, value=f'Comentarios del evaluado: {evaluacion.comentario_evaluado}')
    if evaluacion.comentario_supervisor:
        worksheet.cell(row=54, column=1, value=f'Comentarios del supervisor: {evaluacion.comentario_supervisor}')
    if evaluacion.comentario_gghh:
        worksheet.cell(row=55, column=1, value=f'Comentarios de la gerencia: {evaluacion.comentario_gghh}')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="resultados_operativo_{evaluacion.evaluado.ficha}_{evaluacion.periodo.fecha_inicio}_{evaluacion.periodo.fecha_fin}.xlsx"'
    return response

def fill_resultado_apoyo(evaluacion):
    workbook = openpyxl.load_workbook('core/reportes/bases/resultados-apoyo.xlsx')
    worksheet = workbook.active

    worksheet.cell(row=6, column=1, value=f"EVALUACIÓN DE DESEMPEÑO")
    worksheet.cell(row=6, column=1, value=f"PERÍODO DE EVALUACIÓN DEL {evaluacion.periodo.fecha_inicio.strftime('%d/%m/%Y')} AL {evaluacion.periodo.fecha_fin.strftime('%d/%m/%Y')}")

    worksheet.cell(row=8, column=2, value=evaluacion.evaluado.user.get_full_name())
    worksheet.cell(row=8, column=18, value=evaluacion.evaluado.ficha)
    worksheet.cell(row=9, column=2, value=evaluacion.evaluado.cargo.nombre)

    worksheet.cell(row=14, column=11, value=evaluacion.escalafones.get(asignado_por="H").escalafon.nivel.upper())
    
    resultado_instrumento = evaluacion.resultados.get(instrumento__nombre__icontains='Evaluación del Desempeño')
    worksheet.cell(row=13, column=2, value=float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Costos').resultado_final) / float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Costos').seccion.peso) * (float(resultado_instrumento.instrumento.peso) * float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Costos').seccion.peso) / 100))
    worksheet.cell(row=13, column=3, value=float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Cantidad').resultado_final) / float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Cantidad').seccion.peso) * (float(resultado_instrumento.instrumento.peso) * float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Cantidad').seccion.peso) / 100))
    worksheet.cell(row=13, column=4, value=float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Calidad').resultado_final) / float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Calidad').seccion.peso) * (float(resultado_instrumento.instrumento.peso) * float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Calidad').seccion.peso) / 100))
    worksheet.cell(row=13, column=5, value=float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Seguridad').resultado_final) / float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Seguridad').seccion.peso) * (float(resultado_instrumento.instrumento.peso) * float(resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Seguridad').seccion.peso) / 100))
    worksheet.cell(row=13, column=6, value=float(resultado_instrumento.resultado_final))      
    
    resultado_instrumento = evaluacion.resultados.get(instrumento__nombre__icontains='Competencias Técnicas')
    worksheet.cell(row=13, column=7, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Capacidades Operativas').resultado_final)
    worksheet.cell(row=13, column=8, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Capacidades de Negociación y Relación').resultado_final)
    worksheet.cell(row=13, column=9, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Capacidades de Gestión').resultado_final)
    worksheet.cell(row=13, column=10, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Desarrollo').resultado_final)
    worksheet.cell(row=13, column=11, value=resultado_instrumento.resultado_final)

    resultado_instrumento = evaluacion.resultados.get(instrumento__nombre__icontains='Competencias Genérica')
    worksheet.cell(row=13, column=12, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Liderazgo:').resultado_final)
    worksheet.cell(row=13, column=13, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Adaptabilidad:').resultado_final)
    worksheet.cell(row=13, column=14, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Relaciones:').resultado_final)
    worksheet.cell(row=13, column=15, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Realización de Tareas:').resultado_final)
    worksheet.cell(row=13, column=16, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Producción:').resultado_final)
    worksheet.cell(row=13, column=17, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Desarrollo de los demás:').resultado_final)
    worksheet.cell(row=13, column=18, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Desarrollo Personal:').resultado_final)
    worksheet.cell(row=13, column=19, value=resultado_instrumento.resultados_secciones.get(seccion__nombre__icontains='Comunicación:').resultado_final)
    worksheet.cell(row=13, column=20, value=resultado_instrumento.resultado_final)

    logros = evaluacion.logros_y_metas.filter(anadido_por="H", periodo="A")
    formaciones = evaluacion.formaciones.filter(anadido_por="H")

    tecnico_cientificas = formaciones.filter(clasificacion__clasificacion__icontains='Tecno-Cientifica').count()
    supervisoria_formativa = formaciones.filter(clasificacion__clasificacion__icontains='Supervisoria Formativa').count()
    tecnica_administrativa = formaciones.filter(clasificacion__clasificacion__icontains='Tecnica Administrativa').count()
    seguridad_industrial = formaciones.filter(clasificacion__clasificacion__icontains='Seguridad Industrial').count()
    total_cursos = formaciones.count()

    worksheet.cell(row=40, column=3, value=tecnico_cientificas)
    worksheet.cell(row=41, column=3, value=supervisoria_formativa)
    worksheet.cell(row=40, column=12, value=tecnica_administrativa)
    worksheet.cell(row=40, column=12, value=seguridad_industrial)
    worksheet.cell(row=40, column=21, value=total_cursos)
    for i, logro in enumerate(logros, start=17):
        worksheet.cell(row=i, column=1, value=logro.descripcion)
        worksheet.cell(row=i, column=16, value=logro.porc_cumplimiento)

    for i, formacion in enumerate(formaciones, start=31):
        worksheet.cell(row=i, column=1, value=formacion.necesidad_formacion)
        worksheet.cell(row=i, column=5, value=formacion.prioridad)
        worksheet.cell(row=i, column=7, value=formacion.clasificacion.clasificacion)
        col = 13
        if formacion.competencias.filter(nombre__icontains='Capacidades Operativas').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Capacidades de Negociación y Relación').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Capacidades de Gestión').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Desarrollo').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Liderazgo').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Adaptabilidad').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Relaciones').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Realización de Tareas').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Producción').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Desarrollo de los demás').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Desarrollo del Personal').exists():
            worksheet.cell(row=i, column=col, value='X')
        
        col += 1
        if formacion.competencias.filter(nombre__icontains='Comunicación').exists():
            worksheet.cell(row=i, column=col, value='X')

    metas = evaluacion.logros_y_metas.filter(periodo="P", anadido_por="H")[:5]
    for i, meta in enumerate(metas, start=44):
        worksheet.cell(row=i, column=1, value=meta.descripcion)
        worksheet.cell(row=i, column=16, value=meta.prioridad_larga())

    if evaluacion.comentario_evaluado:
        worksheet.cell(row=53, column=1, value=f'Comentarios del evaluado: {evaluacion.comentario_evaluado}')
    if evaluacion.comentario_supervisor:
        worksheet.cell(row=54, column=1, value=f'Comentarios del supervisor: {evaluacion.comentario_supervisor}')
    if evaluacion.comentario_gghh:
        worksheet.cell(row=55, column=1, value=f'Comentarios de la gerencia: {evaluacion.comentario_gghh}')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="resultados_apoyo_{evaluacion.evaluado.ficha}_{evaluacion.periodo.fecha_inicio}_{evaluacion.periodo.fecha_fin}.xlsx"'
    return response
